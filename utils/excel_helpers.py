"""Copy-on-write Excel population, cell mapping, and LibreOffice recalc."""

import json
import logging
import os
import shutil
import subprocess
from dataclasses import dataclass, field
from datetime import datetime, timezone

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class FlatMapping:
    """Maps a data key to a single cell, e.g. {"key": "revenue", "sheet": "Model", "cell": "B5"}."""
    key: str
    sheet: str
    cell: str


@dataclass
class RangeMapping:
    """Maps a data key (list) to a range of cells.

    direction: "vertical" writes down rows, "horizontal" writes across columns.
    start_cell: top-left cell of the range, e.g. "B10".
    """
    key: str
    sheet: str
    start_cell: str
    direction: str = "vertical"


@dataclass
class CellMappingConfig:
    schema_version: str
    input_flat: list[FlatMapping] = field(default_factory=list)
    input_ranges: list[RangeMapping] = field(default_factory=list)
    output_flat: list[FlatMapping] = field(default_factory=list)


def load_cell_mapping(path: str) -> CellMappingConfig:
    """Parse cell_mapping.json into a CellMappingConfig."""
    with open(path) as f:
        raw = json.load(f)

    config = CellMappingConfig(schema_version=raw.get("schema_version", "1.0"))

    for item in raw.get("inputs", {}).get("flat", []):
        config.input_flat.append(FlatMapping(**item))

    for item in raw.get("inputs", {}).get("ranges", []):
        config.input_ranges.append(RangeMapping(**item))

    for item in raw.get("outputs", {}).get("flat", []):
        config.output_flat.append(FlatMapping(**item))

    return config


def copy_base_model(base_path: str, output_dir: str, timestamp: str) -> str:
    """Copy base_model.xlsx to outputs/ with a timestamp. Returns the new path."""
    if not os.path.isfile(base_path):
        raise FileNotFoundError(f"Base model not found: {base_path}")
    os.makedirs(output_dir, exist_ok=True)
    dest = os.path.join(output_dir, f"{timestamp}_model.xlsx")
    shutil.copy2(base_path, dest)
    logger.info("Copied base model to %s", dest)
    return dest


def validate_keys(mapping: CellMappingConfig, stage2b_data: dict) -> list[str]:
    """Check that every mapped input key exists in the Stage 2b data.

    Returns a list of missing keys (empty list means all good).
    """
    required_keys = set()
    for m in mapping.input_flat:
        required_keys.add(m.key)
    for m in mapping.input_ranges:
        required_keys.add(m.key)

    available_keys = set(stage2b_data.keys())
    missing = sorted(required_keys - available_keys)

    if missing:
        logger.error(
            "Key validation failed.\n  Missing keys: %s\n  Available keys: %s",
            missing,
            sorted(available_keys),
        )
    return missing


def _col_letter_to_number(col: str) -> int:
    """Convert Excel column letter(s) to 1-based number. A=1, Z=26, AA=27."""
    result = 0
    for c in col.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def _number_to_col_letter(n: int) -> str:
    """Convert 1-based column number to letter(s)."""
    result = []
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def _parse_cell_ref(cell: str) -> tuple[str, int]:
    """Split 'B10' into ('B', 10)."""
    col = ""
    row = ""
    for ch in cell:
        if ch.isalpha():
            col += ch
        else:
            row += ch
    return col, int(row)


def write_values_to_excel(
    path: str,
    mapping: CellMappingConfig,
    data: dict,
) -> None:
    """Write Stage 2b data into the Excel copy using the cell mapping."""
    wb = openpyxl.load_workbook(path)

    # Flat mappings
    for m in mapping.input_flat:
        value = data.get(m.key)
        if value is None:
            continue
        ws = wb[m.sheet]
        ws[m.cell] = value
        logger.debug("Wrote %s = %s to %s!%s", m.key, value, m.sheet, m.cell)

    # Range mappings
    for m in mapping.input_ranges:
        values = data.get(m.key)
        if not isinstance(values, list):
            continue
        ws = wb[m.sheet]
        col_str, start_row = _parse_cell_ref(m.start_cell)
        col_num = _col_letter_to_number(col_str)

        for i, val in enumerate(values):
            if m.direction == "vertical":
                cell_ref = f"{col_str}{start_row + i}"
            else:  # horizontal
                cell_ref = f"{_number_to_col_letter(col_num + i)}{start_row}"
            ws[cell_ref] = val
            logger.debug("Wrote %s[%d] = %s to %s!%s", m.key, i, val, m.sheet, cell_ref)

    wb.save(path)
    logger.info("Excel values written to %s", path)


def recalculate_formulas(path: str) -> bool:
    """Attempt to recalculate formulas via LibreOffice headless.

    Returns True if recalc succeeded, False otherwise.
    """
    try:
        output_dir = os.path.dirname(path)
        result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                output_dir,
                path,
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )
        if result.returncode == 0:
            logger.info("LibreOffice recalc succeeded for %s", path)
            return True
        else:
            logger.warning("LibreOffice recalc failed: %s", result.stderr)
            return False
    except FileNotFoundError:
        logger.warning("LibreOffice not found — skipping formula recalculation")
        return False
    except subprocess.TimeoutExpired:
        logger.warning("LibreOffice recalc timed out")
        return False


def read_output_cells(path: str, mapping: CellMappingConfig) -> dict:
    """Read output cells from the Excel file.

    Uses data_only=True so cached values are returned instead of formulas.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    outputs = {}
    for m in mapping.output_flat:
        ws = wb[m.sheet]
        outputs[m.key] = ws[m.cell].value
        logger.debug("Read output %s = %s from %s!%s", m.key, outputs[m.key], m.sheet, m.cell)
    return outputs
